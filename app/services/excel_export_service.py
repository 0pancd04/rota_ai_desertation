import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Any
import io
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ExcelExportService:
    def __init__(self):
        # Define color schemes for highlighting
        self.colors = {
            'header': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'header_font': Font(color='FFFFFF', bold=True),
            'alternate_row': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
            'priority_high': PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),
            'priority_medium': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
            'priority_low': PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid'),
            'unassigned': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
            'assigned': PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid'),
            'available': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
            'unavailable': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def export_assignments_data(self, assignments: List[Dict], patients: List[Dict], employees: List[Dict]) -> bytes:
        """
        Export assignments data to Excel with three sheets:
        1. Assignments - All current assignments
        2. Patients - All patients with assignment status
        3. Employees - All employees with availability and workload
        """
        try:
            logger.info(f"Starting Excel export - Assignments: {len(assignments)}, Patients: {len(patients)}, Employees: {len(employees)}")
            
            # Validate input data
            if not isinstance(assignments, list):
                logger.warning("Assignments is not a list, converting to empty list")
                assignments = []
            if not isinstance(patients, list):
                logger.warning("Patients is not a list, converting to empty list")
                patients = []
            if not isinstance(employees, list):
                logger.warning("Employees is not a list, converting to empty list")
                employees = []
            
            logger.info("Creating Excel workbook...")
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            logger.info("Creating assignments sheet...")
            # Create assignments sheet
            self._create_assignments_sheet(wb, assignments)
            
            logger.info("Creating patients sheet...")
            # Create patients sheet
            self._create_patients_sheet(wb, patients, assignments)
            
            logger.info("Creating employees sheet...")
            # Create employees sheet
            self._create_employees_sheet(wb, employees, assignments)
            
            logger.info("Saving workbook to bytes...")
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            result = output.getvalue()
            logger.info(f"Excel export completed successfully. Size: {len(result)} bytes")
            
            return result
            
        except Exception as e:
            logger.error(f"Error exporting Excel data: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise Exception(f"Failed to export Excel data: {str(e)}")

    def _create_assignments_sheet(self, wb: openpyxl.Workbook, assignments: List[Dict]):
        """Create the assignments sheet with detailed assignment information"""
        ws = wb.create_sheet("Assignments")
        
        # Define headers
        headers = [
            "Assignment ID", "Employee ID", "Employee Name", "Patient ID", "Patient Name",
            "Service Type", "Assigned Time", "Start Time", "End Time", "Duration (mins)",
            "Travel Time (mins)", "Priority Score", "Assignment Reason", "Status"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.colors['header_font']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.colors['border']
        
        # Add data rows
        if assignments:
            for row, assignment in enumerate(assignments, 2):
                if not isinstance(assignment, dict):
                    continue
                    
                # Determine status based on assignment data
                status = "Active" if assignment.get('assigned_time') else "Pending"
                
                row_data = [
                    f"ASG{row-1:04d}",
                    assignment.get('employee_id', ''),
                    assignment.get('employee_name', ''),
                    assignment.get('patient_id', ''),
                    assignment.get('patient_name', ''),
                    assignment.get('service_type', ''),
                    assignment.get('assigned_time', ''),
                    assignment.get('start_time', ''),
                    assignment.get('end_time', ''),
                    assignment.get('duration', 0),
                    assignment.get('travel_time', 0),
                    assignment.get('priority_score', 0),
                    assignment.get('reasoning', ''),
                    status
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.colors['border']
                    
                    # Apply conditional formatting
                    if col == 12:  # Priority Score column
                        priority = assignment.get('priority_score', 0)
                        if priority >= 8:
                            cell.fill = self.colors['priority_high']
                        elif priority >= 6:
                            cell.fill = self.colors['priority_medium']
                        else:
                            cell.fill = self.colors['priority_low']
                    
                    # Alternate row colors
                    if row % 2 == 0:
                        cell.fill = self.colors['alternate_row']
        else:
            # Add a row indicating no assignments
            ws.cell(row=2, column=1, value="No assignments found")
            ws.cell(row=2, column=2, value="").fill = self.colors['alternate_row']
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_patients_sheet(self, wb: openpyxl.Workbook, patients: List[Dict], assignments: List[Dict]):
        """Create the patients sheet with patient information and assignment status"""
        ws = wb.create_sheet("Patients")
        
        # Define headers
        headers = [
            "Patient ID", "Patient Name", "Address", "Post Code", "Gender", "Ethnicity", "Religion",
            "Required Support", "Required Hours", "Additional Requirements", "Illness",
            "Contact Number", "Requires Medication", "Emergency Contact", "Emergency Relation",
            "Language Preference", "Notes", "Assignment Status", "Assigned Employee", "Service Type"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.colors['header_font']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.colors['border']
        
        # Create assignment lookup
        assignment_lookup = {}
        for assignment in assignments:
            if isinstance(assignment, dict):
                patient_id = assignment.get('patient_id')
                if patient_id:
                    assignment_lookup[patient_id] = assignment
        
        # Add data rows
        if patients:
            for row, patient in enumerate(patients, 2):
                if not isinstance(patient, dict):
                    continue
                    
                # Check assignment status
                assignment = assignment_lookup.get(patient.get('patient_id', ''))
                assignment_status = "Assigned" if assignment else "Unassigned"
                assigned_employee = assignment.get('employee_name', '') if assignment else 'N/A'
                service_type = assignment.get('service_type', '') if assignment else 'N/A'
                
                row_data = [
                    patient.get('patient_id', ''),
                    patient.get('patient_name', ''),
                    patient.get('address', ''),
                    patient.get('postcode', ''),
                    patient.get('gender', ''),
                    patient.get('ethnicity', ''),
                    patient.get('religion', ''),
                    patient.get('required_support', ''),
                    patient.get('required_hours_of_support', 0),
                    patient.get('additional_requirements', ''),
                    patient.get('illness', ''),
                    patient.get('contact_number', ''),
                    patient.get('requires_medication', ''),
                    patient.get('emergency_contact', ''),
                    patient.get('emergency_relation', ''),
                    patient.get('language_preference', ''),
                    patient.get('notes', ''),
                    assignment_status,
                    assigned_employee,
                    service_type
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.colors['border']
                    
                    # Apply conditional formatting for assignment status
                    if col == 18:  # Assignment Status column
                        if assignment_status == "Assigned":
                            cell.fill = self.colors['assigned']
                        else:
                            cell.fill = self.colors['unassigned']
                    
                    # Alternate row colors
                    if row % 2 == 0:
                        cell.fill = self.colors['alternate_row']
        else:
            # Add a row indicating no patients
            ws.cell(row=2, column=1, value="No patients found")
            ws.cell(row=2, column=2, value="").fill = self.colors['alternate_row']
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_employees_sheet(self, wb: openpyxl.Workbook, employees: List[Dict], assignments: List[Dict]):
        """Create the employees sheet with employee information, availability, and workload"""
        ws = wb.create_sheet("Employees")
        
        # Define headers
        headers = [
            "Employee ID", "Name", "Address", "Post Code", "Gender", "Ethnicity", "Religion",
            "Transport Mode", "Qualification", "Languages Spoken", "Certificate Expiry",
            "Earliest Start", "Latest End", "Shifts", "Contact Number", "Notes",
            "Current Assignments", "Max Patients/Day", "Workload %", "Availability Status",
            "Assigned Patients", "Total Working Hours", "Total Travel Time"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.colors['header_font']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.colors['border']
        
        # Create assignment lookup for employees
        employee_assignments = {}
        for assignment in assignments:
            if isinstance(assignment, dict):
                employee_id = assignment.get('employee_id')
                if employee_id:
                    if employee_id not in employee_assignments:
                        employee_assignments[employee_id] = []
                    employee_assignments[employee_id].append(assignment)
        
        # Add data rows
        if employees:
            for row, employee in enumerate(employees, 2):
                if not isinstance(employee, dict):
                    continue
                    
                employee_id = employee.get('employee_id', '')
                current_assignments = employee_assignments.get(employee_id, [])
                current_count = len(current_assignments)
                max_patients = employee.get('max_patients_per_day', 8)
                workload_percentage = (current_count / max_patients) * 100 if max_patients > 0 else 0
                
                # Determine availability status
                if workload_percentage >= 100:
                    availability_status = "Fully Booked"
                elif workload_percentage >= 80:
                    availability_status = "Limited Availability"
                elif workload_percentage >= 50:
                    availability_status = "Moderate Availability"
                else:
                    availability_status = "Available"
                
                # Calculate total working hours and travel time
                total_working_hours = sum(assignment.get('duration', 0) for assignment in current_assignments) / 60
                total_travel_time = sum(assignment.get('travel_time', 0) for assignment in current_assignments)
                
                # Get assigned patient names
                assigned_patients = ', '.join([a.get('patient_name', '') for a in current_assignments]) if current_assignments else 'None'
                
                row_data = [
                    employee.get('employee_id', ''),
                    employee.get('name', ''),
                    employee.get('address', ''),
                    employee.get('postcode', ''),
                    employee.get('gender', ''),
                    employee.get('ethnicity', ''),
                    employee.get('religion', ''),
                    employee.get('transport_mode', ''),
                    employee.get('qualification', ''),
                    employee.get('language_spoken', ''),
                    employee.get('certificate_expiry_date', ''),
                    employee.get('earliest_start', ''),
                    employee.get('latest_end', ''),
                    employee.get('shifts', ''),
                    employee.get('contact_number', ''),
                    employee.get('notes', ''),
                    current_count,
                    max_patients,
                    f"{workload_percentage:.1f}%",
                    availability_status,
                    assigned_patients,
                    f"{total_working_hours:.1f}",
                    f"{total_travel_time} mins"
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.colors['border']
                    
                    # Apply conditional formatting for workload
                    if col == 19:  # Workload % column
                        if workload_percentage >= 100:
                            cell.fill = self.colors['unavailable']
                        elif workload_percentage >= 80:
                            cell.fill = self.colors['priority_medium']
                        else:
                            cell.fill = self.colors['available']
                    
                    # Apply conditional formatting for availability status
                    if col == 20:  # Availability Status column
                        if availability_status == "Fully Booked":
                            cell.fill = self.colors['unavailable']
                        elif availability_status == "Limited Availability":
                            cell.fill = self.colors['priority_medium']
                        elif availability_status == "Moderate Availability":
                            cell.fill = self.colors['priority_low']
                        else:
                            cell.fill = self.colors['available']
                    
                    # Alternate row colors
                    if row % 2 == 0:
                        cell.fill = self.colors['alternate_row']
        else:
            # Add a row indicating no employees
            ws.cell(row=2, column=1, value="No employees found")
            ws.cell(row=2, column=2, value="").fill = self.colors['alternate_row']
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _auto_adjust_columns(self, ws: Worksheet):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
